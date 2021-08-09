from tkinter import *
import subprocess
import platform
import openpyxl
from Classes import Machine
from Classes import Takash

# colors:
BLACK1 = '#252525'
BLACK2 = '#404040'
RED = '#c00002'
GREEN = '#82c829'
WHITE = '#FFFFFF'

# files name
xl_file = "input.xlsx"

# measures
WIDTH = 900
HEIGHT = 600
MAIN_WINDOW_SIZE = str(WIDTH)+'x'+str(HEIGHT)

right_WIDTH = 280
right_HEIGHT = 500

SERVER_FRAME_WIDTH = 280
SERVER_FRAME_HEIGHT = 120

TAKASH_FRAME_WIDTH = 280
TAKASH_FRAME_HEIGHT = 240

TAKASH_WIDTH = 80
TAKASH_HEIGHT = 150

frames = []

def get_headquarters_name(): ### temp for now
    return("bla bla 36")

def send_one_ping(takash_ip):

    response = send_ping(takash_ip)
    print(takash_ip)
    if response:
        print("good")
    else:
        print("bad")

def send_ping(current_ip_address):
    try:
        output = subprocess.check_output("ping -{} 1 {}".format('n' if platform.system().lower(
        ) == "windows" else 'c', current_ip_address), shell=True, universal_newlines=True)
        if 'unreachable' in output:
            return False
        else:
            return True
    except Exception:
        return False

def add_frame(frame_to_add_to, title_name, array_of_machines, takash_ip):
    frame = Frame(frame_to_add_to, bg=BLACK2, width=TAKASH_WIDTH, height=TAKASH_HEIGHT)
    frame.pack(side=LEFT, padx=5, pady=5)

    name_label = Label(frame, text=title_name, bg=BLACK2, fg=WHITE, font='Helvetica 10 bold')
    name_label.grid(row=0, column=0, sticky=EW)

    for index,machine in enumerate(array_of_machines):
        button = Button(frame, text=machine , bg=GREEN ,command= lambda: send_one_ping(takash_ip), fg=WHITE, font='Helvetica 8 bold')
        button.grid(row=index+1, column=0, sticky=EW)

def create_top_title(title_name, frame_to_add_to, buttons):

    upper_frame = Frame(frame_to_add_to, bg=BLACK2)
    upper_frame.pack(fill=X, pady=2, padx=5)

    if buttons == True:
        setting_button = Button(upper_frame, text=" הגדרות", bg=BLACK1, fg=WHITE, font='Helvetica 8 bold', width = 6)
        setting_button.grid(row=0, column=0, sticky=W)
        statistics_button = Button(upper_frame, text="סטיסטיקה", bg=BLACK1, fg=WHITE, font='Helvetica 8 bold', width = 6)
        statistics_button.grid(row=1, column=0, sticky=W )

    title_bottom = Label(upper_frame, text=title_name, bg=BLACK2, fg=WHITE, font='Helvetica 18 bold')
    title_bottom.grid(row=0, column=1, rowspan=2)
    upper_frame.grid_rowconfigure(1, weight=1)
    upper_frame.grid_columnconfigure(1, weight=1)


    span = Frame(frame_to_add_to, bg=BLACK1, height=2)
    span.pack(fill=X, padx=15)

def create_left_side(main_frame):
    left_frame = Frame(main_frame, bg=BLACK2, width=280, height=500)
    left_frame.pack(fill=BOTH, expand=True, side=LEFT, pady=10, padx=10)
    left_frame.pack_propagate(0)

    create_top_title("ליבות ואתרים ניידים", left_frame, False)

    return left_frame

def create_right_side(main_frame):
    right_frame = Frame(main_frame, bg=BLACK2, width=right_WIDTH, height=right_HEIGHT)
    right_frame.pack(fill=BOTH, expand=True, side=RIGHT)
    right_frame.pack_propagate(0)

    headquarters_name = get_headquarters_name()
    create_top_title(headquarters_name, right_frame, True)

    server_frame = Frame(right_frame, bg=RED, width=SERVER_FRAME_WIDTH, height=SERVER_FRAME_HEIGHT)
    server_frame.pack(fill=X, expand=True)
    server_frame.pack_propagate(0)

    span = Frame(right_frame, bg=BLACK1, height=2)
    span.pack(fill=X, padx=15)

    takash_frame = Frame(right_frame, bg=GREEN, width=TAKASH_FRAME_WIDTH, height=TAKASH_FRAME_HEIGHT)
    takash_frame.pack(fill=X, expand=True)
    takash_frame.pack_propagate(0)

    span = Frame(right_frame, bg=BLACK1, height=2)
    span.pack(fill=X, padx=15)

    carriage_frame = Frame(right_frame, bg=GREEN, width=TAKASH_FRAME_WIDTH, height=TAKASH_FRAME_HEIGHT)
    carriage_frame.pack(fill=X, expand=True)
    carriage_frame.pack_propagate(0)

    return server_frame, takash_frame, carriage_frame

def create_middle(main_frame):
    middle_top_frame = Frame(main_frame, bg = BLACK2,width=280, height=250)
    middle_top_frame.pack(fill=BOTH, expand=True, side=TOP, pady=10, padx=5)
    middle_top_frame.pack_propagate(0)

    headquarters_name = get_headquarters_name()
    create_top_title(headquarters_name, middle_top_frame, True)


    middle_bottom_frame = Frame(main_frame, bg = BLACK2, width=280, height=250)
    middle_bottom_frame.pack(fill=BOTH, expand=True, side=BOTTOM, pady=10, padx=5)
    middle_bottom_frame.pack_propagate(0)

    create_top_title("סטיסטיקה וזה", middle_bottom_frame, False)

    return middle_top_frame, middle_bottom_frame

def creating_main_window(main_window):

    main_window.title('Pingger')
    main_window.geometry(MAIN_WINDOW_SIZE)

    main_frame = Frame(main_window, bg = BLACK1)
    main_frame.pack(fill=BOTH, expand=True)

    left_frame = create_left_side(main_frame)
    server_frame, takash_frame, carriage_frame = create_right_side(main_frame)
    middle_top_frame, middle_bottom_frame = create_middle(main_frame)

    return left_frame, middle_top_frame, middle_bottom_frame, server_frame, takash_frame, carriage_frame

def insert_data_to_ui(frame):
    #first insert the right side of the UI
    #open the xl file in the right sheet and take the right data
    path = xl_file
    sheet_name = "mapping"
    dict_of_takashes = get_data_from_file(path, sheet_name)

    temp_machine_name_list = []
    temp_ip_list = []
    for key in dict_of_takashes:
        for machine in dict_of_takashes[key].machine_list:
            print(machine.who_am_i)
            print(machine.ip)

    for key in dict_of_takashes:
        for machine in dict_of_takashes[key].machine_list:
            temp_machine_name_list.append(machine.who_am_i)
            temp_ip_list.append(machine.ip)
        '''add_frame(frame_to_add_to, title_name, array_of_machines, takash_ip)'''
        add_frame(frame, key , temp_machine_name_list, temp_ip_list)
        temp_machine_name_list=[]
        temp_ip_list=[]



def get_data_from_file(path,sheet):
    print("start fetching data from file: " + path + "at sheet: " + sheet)

    # To open the workbook, workbook object is created
    workbook_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object from the active attribute
    sheet_obj = workbook_obj.active

    current_takash = sheet_obj.cell(row=2, column=1).value
    takash_with_machine = [] #tuple of (takash name, machine name)
    # Note: The first row or column integer is 1, not 0.

    all_machines = {}
    all_takash = {}
    machines_to_add = 0
    first_time_flag = True
    current_row = []
    for row in sheet_obj.iter_rows(min_row=2):
         cell_count = 1
         for cell in row:
             if cell_count == 1:
                 name_of_takash = cell.value
                 if current_takash != cell.value:
                     current_takash = cell.value
                     machines_to_add = 0
                 machines_to_add += 1
                 current_row = []
             else:
                current_row.append(cell.value)
             cell_count += 1

         all_machines[current_row[2]] = Machine(current_row[2],current_row[1],current_row[0], current_row[3], False)
         takash_with_machine.append([name_of_takash, all_machines[current_row[2]]])

    takash_with_machine.pop()
    # print(takash_with_machine)
    last_takash_name = ""
    for tuple in takash_with_machine: #tuple of (takash name, machine name)
        if last_takash_name != tuple[0]:
            all_takash[tuple[0]] = Takash(tuple[0])
            last_takash_name = tuple[0]
        all_takash[tuple[0]].add_machine(tuple[1])

    return all_takash


if __name__ == '__main__':

    # run the algo and show the object
    # list_of_all_addresses = read_old_data()

    main_window = Tk()  # create root window
    # main_window.iconbitmap("icon_image.ico")

    

    '''left_frame, middle_top_frame, middle_bottom_frame, server_frame, takash_frame, carriage_frame = creating_main_window(main_window)

    insert_data_to_ui(server_frame)'''

    # arr = ["david", "elkabas", "google"]
    # arr2 = ["8.8.8.8", "28.82.4.24", "123.2.3.4"]
    # for i in range(0, 3):
    #     add_frame(server_frame, "תקש 123", arr, arr2[i])

    #insert_data(right_frame.)
    # print("hello")

    main_window.mainloop()
