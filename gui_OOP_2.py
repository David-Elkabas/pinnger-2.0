import tkinter as tk
from tkinter import ttk
import tkinter.font as tkFont
import subprocess
import json
# import treading
import concurrent.futures
import platform
import openpyxl
from Classes import Machine
from Classes import Takash
from datetime import datetime
import winsound
import time

# Opening JSON file
try:
    file = open("setting.json", encoding="utf8")
except Exception as e:
    print(e)
    # returns JSON object as
    # a dictionary
setting_data = json.load(file)
# colors:
BLACK1 = setting_data["colors"]["BLACK1"]
BLACK2 = setting_data["colors"]["BLACK2"]
RED = setting_data["colors"]["RED"]
GREEN = setting_data["colors"]["GREEN"]
WHITE = setting_data["colors"]["WHITE"]
GREY = setting_data["colors"]["GREY"]

'''change MORE_data to True if you wish to add additional data from 'more' sheet'''
MORE_DATA = setting_data["MORE_TAKASHIM"]

'''change READ_FROM_JSON to True if you wish read the data from a json file'''
READ_FROM_JSON = setting_data["read_from_json"]

IP_SOURCE_ONE = setting_data["ip_source_1"]
IP_SOURCE_TWO = setting_data["ip_source_2"]

UPDATE_UI_TIME_MSEC = setting_data["UPDATE_UI_TIME_PER_AREA_IN_MSEC"]
SEND_PING_TIME_MSEC = setting_data["PING_SEND_TIME_IN_MSEC"]

# files name
xl_file = setting_data["xl_file_name"]
json_file = setting_data["json_file_name"]
if READ_FROM_JSON:
    read_from_file = json_file
else:
    read_from_file = xl_file

# measures

FONT_SIZE = setting_data["font_size"]
WIDTH = 900
HEIGHT = 840
MAIN_WINDOW_SIZE = str(WIDTH) + 'x' + str(HEIGHT)

right_WIDTH = 400
right_HEIGHT = HEIGHT

SERVER_FRAME_WIDTH = 280
SERVER_FRAME_HEIGHT = 120

TAKASH_FRAME_WIDTH = 280
TAKASH_FRAME_HEIGHT = 240

TAKASH_WIDTH = 80
TAKASH_HEIGHT = 150

# def flickering_color(button, color):
#     for i in range(0,2):
#         button.configure(bg=BLACK1)
#         time.sleep(0.1)
#         button.configure(bg=GREEN)
#         time.sleep(0.1)


# TODO - add Fonts to EVERYTHING
class Fonts:
    def __init__(self):
        self.customFont_small = tkFont.Font(family="Helvetica", size=8)
        self.customFont = tkFont.Font(family="Helvetica", size=10)
        self.customFont_big = tkFont.Font(family="Helvetica", size=12)
        self.customFont_huge = tkFont.Font(family="Helvetica", size=18)

    def OnBigger(self):
        '''Make the font 2 points bigger'''
        size = self.customFont['size']
        self.customFont.configure(size=size+2)

    def OnSmaller(self):
        '''Make the font 2 points smaller'''
        size = self.customFont['size']
        self.customFont.configure(size=size-2)


class Program(tk.Tk, Fonts):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        Fonts.__init__(self)
        self.update_state = 0
        if MORE_DATA:
            w = WIDTH
            h = HEIGHT
        else:
            w = right_WIDTH
            h = right_HEIGHT
        self.main_container = tk.Frame(self, bg=BLACK2, width=w, height=h)
        self.main_container.pack(fill=tk.BOTH, expand=True, side=tk.RIGHT)
        self.main_container.pack_propagate(0)

        self.right_container = tk.Frame(self.main_container, bg=BLACK2, width=right_WIDTH, height=right_HEIGHT)
        self.right_container.pack(fill=tk.BOTH, expand=True, side=tk.RIGHT)
        self.right_container.pack_propagate(0)

        self.create_right_side()

        if MORE_DATA:
            self.left_container = tk.Frame(self.main_container, bg=BLACK2, width=right_WIDTH, height=right_HEIGHT)
            self.left_container.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
            self.left_container.pack_propagate(0)
            self.create_left_side()
        # self.container.grid_rowconfigure(0, weight=1)
        # self.container.grid_columnconfigure(0, weight=1)

    def create_left_side(self):
        self.headquarters_name_left = "כלים נוספים לבדיקה"
        self.top_frame_left = Top_title(self.left_container, self.headquarters_name_left, False)
        self.top_frame_left.pack(fill=tk.X)

        self.additional_frame = Servers_frame(self.left_container, self)
        self.additional_frame.pack()

    def create_right_side(self):
        self.headquarters_name = self.get_headquarters_name()
        self.top_frame = Top_title(self.right_container, self.headquarters_name, True)
        self.top_frame.pack(fill=tk.X)

        self.server_frame = Servers_frame(self.right_container, self)
        self.server_frame.pack()

        span = tk.Frame(self.right_container, bg=BLACK1, height=2)
        span.pack(fill=tk.X, pady=5, padx=10)

        self.takash_frame = Servers_frame(self.right_container, self)
        self.takash_frame.pack()

        span = tk.Frame(self.right_container, bg=BLACK1, height=2)
        span.pack(fill=tk.X, pady=5, padx=10)

        self.carriage_frame = Servers_frame(self.right_container, self)
        self.carriage_frame.pack()

    def get_headquarters_name(self):  ### temp for now

        if READ_FROM_JSON:
            try:
                file = open(read_from_file, encoding="utf8")
            except Exception as e:
                print(e)
                # returns JSON object as
                # a dictionary
            data_from_file = json.load(file)
            return data_from_file["name"]
        else:
            workbook_obj = openpyxl.load_workbook(read_from_file)
            sheet_obj = workbook_obj["שרתים"]
            current_takash = sheet_obj.cell(row=1, column=8).value
        return current_takash

    def update_data_in_ui(self):

        update_state_place_name = "לקראת עדכון אזור שרתים"
        self.top_frame.update_state_label.configure(text=update_state_place_name)
        if self.update_state == 0:
            self.update_state = 1
            self.after(UPDATE_UI_TIME_MSEC, self.update_server_frame_in_ui)

    def update_server_frame_in_ui(self):
        update_state_place_name = "לקראת עדכון אזור תאי קשר"
        self.top_frame.update_state_label.configure(text=update_state_place_name)
        x = insert_data_to_ui(app.server_frame, "שרתים", False)
        if self.update_state == 1:
            self.update_state = 2
            self.after(UPDATE_UI_TIME_MSEC, self.update_takash_frame_in_ui)

    def update_takash_frame_in_ui(self):
        update_state_place_name = "לקראת עדכון אזור קרונות"
        self.top_frame.update_state_label.configure(text=update_state_place_name)
        y = insert_data_to_ui(app.takash_frame, "תא קשר", False)
        if self.update_state == 2:
            if MORE_DATA == True:
                self.update_state = 3
                self.after(UPDATE_UI_TIME_MSEC, self.update_MORE_DATA_frame_in_ui)
            else:
                self.update_state = 4
                self.after(UPDATE_UI_TIME_MSEC, self.update_carriage_frame_in_ui)

    def update_MORE_DATA_frame_in_ui(self):

        update_state_place_name = "לקראת עדכון אזור מידע נוסף"
        self.top_frame.update_state_label.configure(text=update_state_place_name)
        w = insert_data_to_ui(app.additional_frame, "לבדוק בנוסף", False)
        if self.update_state == 3:
            self.update_state = 4
            self.after(UPDATE_UI_TIME_MSEC, self.update_carriage_frame_in_ui)

    def update_carriage_frame_in_ui(self):
        update_state_place_name = "לקראת עדכון"
        self.top_frame.update_state_label.configure(text=update_state_place_name)
        z = insert_data_to_ui(app.carriage_frame, "קרונות", True)
        if self.update_state == 4:
            self.update_state = 0
            self.after(UPDATE_UI_TIME_MSEC, self.update_data_in_ui)


class Servers_frame(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, bg=BLACK2)


class Top_title(tk.Frame, Fonts):
    def __init__(self, parent, title, need_buttons):
        tk.Frame.__init__(self, parent, bg=BLACK2)
        Fonts.__init__(self)
        self.update_state = 0
        self.upper_frame = tk.Frame(self, bg=BLACK2)
        self.upper_frame.pack(fill=tk.X)

        if need_buttons == True:
            self.setting_button = tk.Button(self.upper_frame, text=" הגדרות", bg=BLACK1, fg=WHITE,
                                            font=self.customFont_small, width=10)
            self.setting_button.bind("<Button>", lambda e : NewWindow(parent))
            self.setting_button.grid(row=0, column=0, sticky=tk.W, padx=5)

            self.statistics_button = tk.Button(self.upper_frame, text="סטיסטיקה", bg=BLACK1, fg=WHITE,
                                               font=self.customFont_small, width=10)
            self.statistics_button.grid(row=1, column=0, sticky=tk.W, padx=5)

            bigger = tk.Button(self.upper_frame, text="+", command=self.OnBigger)
            bigger.grid(row=0, column=1, sticky=tk.W)
            smaller = tk.Button(self.upper_frame, text="- ", command=self.OnSmaller)
            smaller.grid(row=1, column=1, sticky=tk.W)

        self.title_top = tk.Label(self.upper_frame, text=title, bg=BLACK2, fg=WHITE, font=self.customFont_huge)
        self.title_top.grid(row=0, column=1, rowspan=2)
        if need_buttons:
            self.update_state_label = tk.Label(self.upper_frame, text=self.update_state, bg=BLACK2, fg=WHITE,
                                               font=self.customFont, width=18)
            self.update_state_label.grid(row=0, column=2, rowspan=2)
        self.upper_frame.grid_rowconfigure(1, weight=1)
        self.upper_frame.grid_columnconfigure(1, weight=1)

        span = tk.Frame(self, bg=BLACK1, height=2)
        span.pack(fill=tk.X, pady=5, padx=10)




class NewWindow(tk.Toplevel):

    def __init__(self, master=None):
        super().__init__(master=master)
        self.title("New Window")
        self.big_frame = tk.Frame(self, bg=BLACK2)
        self.big_frame.pack()
        # set minimum window size value
        self.minsize(250, 180)
        # set maximum window size value
        self.maxsize(250, 180)
        padding_left = tk.Label(self.big_frame, font='Helvetica 12 bold', text=" לא עובד בנתיים", bg=BLACK2, fg="red")
        padding_left.grid(row=0, column=0, sticky=tk.W)
        title_label = tk.Label(self.big_frame, bg=BLACK2, font='Helvetica 10 bold',
                               text="אנא בחר את קצב שליחת" + "\n" + "הפינג ואת זמן הסריקה")
        title_label.grid(row=0, column=1, sticky=tk.E, columnspan=2, pady=20)

        entry_ping = tk.Entry(self.big_frame, width=11)
        entry_ping.grid(row=1, column=0, sticky=tk.E)
        ping_label = tk.Label(self.big_frame, text=":קצב שליחת פינג", width=12, bg=BLACK2, font='Helvetica 10 bold')
        ping_label.grid(row=1, column=1, sticky=tk.E, columnspan=2)

        entry_scan = tk.Entry(self.big_frame, width=11)
        entry_scan.grid(row=2, column=0, sticky=tk.E)
        scan_time_label = tk.Label(self.big_frame, text=":קצב סריקה", font='Helvetica 10 bold', width=12, bg=BLACK2)
        scan_time_label.grid(row=2, column=1, sticky=tk.E, columnspan=2)

        ok_button = tk.Button(self.big_frame, text="אישור", bg=BLACK1, fg=WHITE,
                              font='Helvetica 8 bold', width=10)
        ok_button.grid(row=3, column=0, sticky=tk.W, padx=15, pady=20)

        cancel_button = tk.Button(self.big_frame, text="ביטול", bg=BLACK1, fg=WHITE,
                                  font='Helvetica 8 bold', width=10)
        cancel_button.grid(row=3, column=1, sticky=tk.W, padx=10, pady=20)


class ToolTip(object):

    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 57
        y = y + cy + self.widget.winfo_rooty() + 27
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                         background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()


def Create_tool_tip(widget, text, need_delete):
    # if need_delete:
    #     toolTip.widget = None
    #     toolTip.tipwindow = None
    #     toolTip.id = None
    #     toolTip.x = 0
    #     toolTip.y = 0
    toolTip = ToolTip(widget)

    def enter(event):
        toolTip.showtip(text)

    def leave(event):
        toolTip.hidetip()

    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)


def send_one_ping(ip_or_hostname, button, machine):
    response = send_ping(ip_or_hostname)
    print(f'sent ping to {ip_or_hostname}')
    if not response:
        res = send_ping_with_source_ip(ip_or_hostname, IP_SOURCE_ONE)
        if not res:
            res2 = send_ping_with_source_ip(ip_or_hostname, IP_SOURCE_TWO)
            if not res2:
                print("didn't succeeded to send ping\n")
                machine.three_last_working_state.insert(0, False)
                machine.three_last_working_state.pop()
                print(machine.three_last_working_state)
                color = RED
                update_GUI(button, color)
                return

    machine.three_last_working_state.insert(0, True)
    machine.three_last_working_state.pop()
    print(machine.three_last_working_state)
    color = GREEN
    print("Succeeded to send ping \n")

    update_GUI(button, color)


def update_GUI(button, color):
    button.configure(bg=color)


def send_ping_without_GUI_update(ip_or_hostname):
    response = False  # send_ping(ip_or_hostname)
    print(f'sending ping to {ip_or_hostname} ')
    if not response:
        res = send_ping_with_source_ip(ip_or_hostname, IP_SOURCE_ONE)
        if not res:
            res2 = send_ping_with_source_ip(ip_or_hostname, IP_SOURCE_TWO)
            if not res2:
                return False
    return True


def send_ping(current_ip_address):
    try:
        output = subprocess.check_output("ping -{} 1 {} -w {}".format('n' if platform.system().lower(
        ) == "windows" else 'c', current_ip_address, SEND_PING_TIME_MSEC), shell=True, universal_newlines=True)
        if 'unreachable' in output:
            return False
        else:
            return True
    except Exception:
        return False


def send_ping_with_source_ip(ip_address, source_ip):
    try:
        output = subprocess.check_output("ping -{} 1 {} -w 250 -S {}".format('n' if platform.system().lower(
        ) == "windows" else 'c', ip_address, source_ip), shell=True, universal_newlines=True)
        if 'unreachable' in output:
            return False
        else:
            return True
    except Exception:
        return False


def get_date():
    now = datetime.now()  # current date and time
    date_time = now.strftime("%d/%m/%Y, %H:%M:%S")

    return date_time


def add_frame_to_ui(frame_to_add_to, title_name, array_of_machines, place, is_carriage):
    if is_carriage:
        padding_x = 0
        label_wraplength = 30
        label_background = GREY
        button_font = 'Helvetica 6 bold'
        row_place = 0
        column_place = place
    else:
        padding_x = 5
        label_wraplength = 0
        label_background = BLACK2
        button_font = 'Helvetica 8 bold'
        row_place = int(place / 3)
        column_place = int(place % 3)

    frame = tk.Frame(frame_to_add_to, bg=BLACK1, width=TAKASH_WIDTH, height=300)
    frame.grid(row=row_place, column=column_place, padx=padding_x, pady=2, sticky=tk.N)

    name_label = tk.Label(frame, text=title_name, bd=2, bg=label_background, fg=WHITE,
                          wraplength=label_wraplength, height=2, font='Helvetica 8 bold')
    name_label.grid(row=0, column=0, sticky=tk.NSEW)
    for index, machine in enumerate(array_of_machines):
        if any(machine.three_last_working_state):
            background_button_color = GREEN
        else:
            background_button_color = RED

        button = tk.Button(frame, text=machine.who_am_i, bg=background_button_color,
                           fg=WHITE, font=button_font)
        if machine.ip == ' ' or machine.ip == '' or machine.ip is None:
            ip_or_hostname = machine.hostname
            text_to_show = 'hostname: '
        else:
            ip_or_hostname = machine.ip
            text_to_show = 'ip: '
        button.configure(command=lambda ip_h=ip_or_hostname, b=button, m=machine: send_one_ping(ip_h, b, m))
        button.grid(row=index + 1, column=0, sticky=tk.EW)

        text_to_add = text_to_show + str(ip_or_hostname + "\n" + 'last time: ' + machine.last_send_time)
        Create_tool_tip(button, text=text_to_add, need_delete=False)


def HE_to_EN_sheet_name(name):
    if name == 'שרתים':
        return 'SERVERS'
    if name == 'תא קשר':
        return 'TAKASHIM'
    if name == 'קרונות':
        return 'CARRIAGES'
    if name == 'לבדוק בנוסף':
        return 'MORE DATA'


def print_status_to_cmd(commend, path, sheet_name=''):
    if commend == 'update':
        print(
            "updating - send ping to all of the takashes ip address and save data in xl file: " + path + " at sheet: " + sheet_name)
    elif commend == 'fetch':
        print("fetching the new data from file: " + path + " at sheet: " + sheet_name + "\n")
    elif commend == 'save':
        print("saved all new data to " + path)


def insert_data_to_ui(frame, sheet, is_carriage):
    path = read_from_file

    sheet_name = HE_to_EN_sheet_name(sheet)
    is_first_time = True
    if READ_FROM_JSON:
        dict_of_machines_in_sheet = get_data_from_json_file(path, sheet_name, is_first_time)
    else:
        dict_of_machines_in_sheet = get_data_from_xl_file(path, sheet, is_first_time)

    print_status_to_cmd('update', path, sheet_name)

    '''send ping to all of the takashes ip address and save data in xl file'''
    scanning_all(dict_of_machines_in_sheet, sheet)

    print_status_to_cmd('fetch', path, sheet_name)
    is_first_time = False
    if READ_FROM_JSON:
        dict_of_takashes, dict_of_machines_in_sheet = get_data_from_json_file(path, sheet_name, is_first_time)
    else:
        dict_of_takashes, dict_of_machines_in_sheet = get_data_from_xl_file(path, sheet, is_first_time)
    temp_machine_name_list = []

    if not is_carriage:
        if len(dict_of_takashes.keys()) > 9:
            print("cant be more than 9 TAKASHIM")
            return False

    place = 0
    for key in dict_of_takashes:
        for machine in dict_of_takashes[key].machine_list:
            temp_machine_name_list.append(machine)

        add_frame_to_ui(frame, key, temp_machine_name_list, place, is_carriage)
        place = place + 1
        temp_machine_name_list = []
    return True


def get_data_from_xl_file(path, sheet, first_time_flag):
    # To open the workbook, workbook object is created
    workbook_obj = openpyxl.load_workbook(path)
    sheet_obj = workbook_obj[sheet]

    current_takash = sheet_obj.cell(row=2, column=1).value
    takash_with_machine = []  # tuple of (takash name, machine name)
    # Note: The first row or column integer is 1, not 0.

    dict_of_machines_in_sheet = {}
    all_takash = {}
    machines_to_add = 0
    exit_loop_flag = False
    current_row = []
    for row in sheet_obj.iter_rows(min_row=2, max_col=6):
        if exit_loop_flag == True:
            break
        else:
            cell_count = 1
            for cell in row:
                if exit_loop_flag == True:
                    break
                else:
                    if cell_count == 1:
                        if cell.value == None or cell.value == 'סוף':
                            exit_loop_flag = True
                        else:
                            name_of_takash = cell.value
                            if current_takash != cell.value:
                                current_takash = cell.value
                                machines_to_add = 0
                            machines_to_add += 1
                            current_row = []
                    else:
                        current_row.append(cell.value)
                    cell_count += 1
        '''after we pull out the data for the selected row we construct the Machine with all the data
        using a dictionary with key=hostname that save all the machines'''
        # Machine - (hostname, ip, who_am_i, last_send_time, is_working, three_last_status)
        dict_of_machines_in_sheet[current_row[2]] = Machine(current_row[2], current_row[1], current_row[0],
                                                            current_row[4], current_row[3],
                                                            [current_row[3], current_row[3], current_row[3]])
        takash_with_machine.append([name_of_takash, dict_of_machines_in_sheet[current_row[2]]])
    takash_with_machine.pop()

    last_takash_name = ""

    if first_time_flag:
        return dict_of_machines_in_sheet

    else:
        for tuple in takash_with_machine:  # tuple of (takash name, machine name)
            if last_takash_name != tuple[0]:
                all_takash[tuple[0]] = Takash(tuple[0])
                last_takash_name = tuple[0]
            all_takash[tuple[0]].add_machine(tuple[1])

        return all_takash, dict_of_machines_in_sheet


def get_data_from_json_file(path, sheet, first_time_flag):
    # Opening JSON file
    try:
        file = open(path, encoding="utf8")
    except Exception as e:
        print(e)
        # returns JSON object as
        # a dictionary
    data_from_file = json.load(file)

    takash_with_machine = []  # tuple of (takash name, machine name)

    dict_of_machines_in_sheet = {}
    all_takash = {}
    current_row = []
    for key in data_from_file[sheet]:
        name_of_takash = key["name"]
        for device in key["devices"]:
            hostname = device["hostname"]
            ip = device["ip"]
            machine_type = device["machine_type"]
            last_update_date = device["last_update_date"]
            status = device["status"]
            three_last_working_state = device["three_last_working_state"]

            '''after we pull out the data from the selected place we construct the Machine class with 
                   all the data using a dictionary with key=hostname that save all the machines in it'''
            # Machine - (hostname, ip, who_am_i, last_send_time, is_working, three_last_status)
            #     dict_of_machines_in_sheet[hostname] = Machine(hostname, ip, machine_type, last_update_date, status,
            #     three_last_working_state)
            dict_of_machines_in_sheet[hostname] = Machine(hostname, ip, machine_type, last_update_date,
                                                          status, three_last_working_state)
            takash_with_machine.append([name_of_takash, dict_of_machines_in_sheet[hostname]])

    last_takash_name = ""

    if first_time_flag:
        return dict_of_machines_in_sheet
    else:
        for tuple in takash_with_machine:  # tuple of (takash name, machine name)
            if last_takash_name != tuple[0]:
                all_takash[tuple[0]] = Takash(tuple[0])
                last_takash_name = tuple[0]
            all_takash[tuple[0]].add_machine(tuple[1])
        return all_takash, dict_of_machines_in_sheet


def scanning_all(dict_of_all_addresses, sheet):
    path = read_from_file
    sheet_obj = ""
    if READ_FROM_JSON:
        sheet_obj = HE_to_EN_sheet_name(sheet)
    else:
        workbook_obj = openpyxl.load_workbook(path)
        sheet_obj = workbook_obj[sheet]

    list_address = []
    for key in dict_of_all_addresses:
        ip_address = dict_of_all_addresses[key].ip
        if ip_address == '' or ip_address == ' ':
            list_address.append(dict_of_all_addresses[key].hostname)
        else:
            list_address.append(ip_address)

    with concurrent.futures.ThreadPoolExecutor() as executor:
        results = executor.map(send_ping_without_GUI_update, list_address)

    i = 0
    results = list(results)
    for key in dict_of_all_addresses:
        dict_of_all_addresses[key].is_working = results[i]
        dict_of_all_addresses[key].three_last_working_state.insert(0, results[i])
        dict_of_all_addresses[key].three_last_working_state.pop()
        time = get_date()
        dict_of_all_addresses[key].last_send_time = time
        i += 1

    for hostname in dict_of_all_addresses:
        if READ_FROM_JSON:
            update_json_file(sheet_obj, hostname, dict_of_all_addresses[hostname].is_working,
                             dict_of_all_addresses[hostname].last_send_time,
                             dict_of_all_addresses[hostname].three_last_working_state)
        else:
            update_xlsx_file(sheet_obj, hostname, dict_of_all_addresses[hostname].is_working,
                             dict_of_all_addresses[hostname].last_send_time,
                             dict_of_all_addresses[hostname].three_last_working_state)

    print_status_to_cmd('save', path)
    if not READ_FROM_JSON:
        workbook_obj.save(path)


def update_json_file(obj_name, hostname, status, time, three_last_working_state):
    try:
        file = open(read_from_file, encoding="utf8")
    except Exception as e:
        print(e)

    # returns JSON object as dictionary
    data = json.load(file)
    file.close()

    for server in data[obj_name]:
        for device in server["devices"]:
            if device["hostname"] == hostname:
                device["status"] = status
                device["last_update_date"] = time
                device["three_last_working_state"] = three_last_working_state

    updated_file = open(read_from_file, "w", encoding="utf8")
    json.dump(data, updated_file, ensure_ascii=False)
    updated_file.close()


def update_xlsx_file(sheet_obj, hostname, status, time, three_last_working_state):
    cell_positions = ''
    time_cell_positions = ''
    for row in sheet_obj.rows:
        element = row[3]
        if element.value == hostname:
            cell_col_position = 'E'
            cell_row_position = element.row
            cell_positions = cell_col_position + str(cell_row_position)
            time_cell_positions = 'F' + str(cell_row_position)

    if status == True:
        sheet_obj[cell_positions] = "עובד"
        sheet_obj[cell_positions].fill = fill = openpyxl.styles.GradientFill(stop=("80ff72", "7ee8fa"))
        # print( sheet_obj[cell_positions].value)

    else:
        sheet_obj[cell_positions] = "לא עובד"
        # print(sheet_obj[cell_positions].value)
        sheet_obj[cell_positions].fill = fill = openpyxl.styles.GradientFill(stop=("fc9842", "fe5f75"))

    sheet_obj[time_cell_positions] = time


if __name__ == '__main__':
    app = Program()

    x = insert_data_to_ui(app.server_frame, "שרתים", False)
    y = insert_data_to_ui(app.takash_frame, "תא קשר", False)
    z = insert_data_to_ui(app.carriage_frame, "קרונות", True)

    '''open another place for optional data'''
    if MORE_DATA == True:
        w = insert_data_to_ui(app.additional_frame, "לבדוק בנוסף", False)

    app.update_data_in_ui()
    app.mainloop()
