import tkinter as tk
from tkinter import ttk
import subprocess
# import treading
import concurrent.futures
import platform
import openpyxl
from Classes import Machine
from Classes import Takash
from datetime import datetime
import time

# colors:
BLACK1 = '#252525'
BLACK2 = '#404040'
RED = '#c00002'
GREEN = '#82c829'
WHITE = '#FFFFFF'
GREY = '#5a5956'

# files name
xl_file = "input.xlsx"

# measures
WIDTH = 900
HEIGHT = 600
MAIN_WINDOW_SIZE = str(WIDTH) + 'x' + str(HEIGHT)

right_WIDTH = 400
right_HEIGHT = 600

SERVER_FRAME_WIDTH = 280
SERVER_FRAME_HEIGHT = 120

TAKASH_FRAME_WIDTH = 280
TAKASH_FRAME_HEIGHT = 240

TAKASH_WIDTH = 80
TAKASH_HEIGHT = 150

# def change_color(frame, color):
#     frame.

# def flickering_color(button, color):
#     for i in range(0,2):
#         button.configure(bg=BLACK1)
#         time.sleep(0.1)
#         button.configure(bg=GREEN)
#         time.sleep(0.1)

def send_one_ping(takash_ip, button, sheet, hostname):
    response = send_ping(takash_ip)
    print(takash_ip)
    if response:
        color=GREEN
        status = True
        print("good")
    else:
        print("bad")
        status = False
        color = RED

    path = xl_file
    workbook_obj = openpyxl.load_workbook(path)
    sheet_obj = workbook_obj[sheet]
    time = get_date()
    update_file(sheet_obj, hostname, status, time)

    workbook_obj.save(path)

    update_GUI(button, color, time, takash_ip)

def update_GUI(button, color, time, ip):
    button.configure(bg=color)
    # text_to_add = 'ip: ' + ip + "\nlast time: " + time
    ''' in the future need to be need_delete = True  '''
    # Create_tool_tip(button, text=text_to_add, need_delete = False)

def send_ping(current_ip_address):
    try:
        output = subprocess.check_output("ping -{} 1 {} -w 200ms".format('n' if platform.system().lower(
        ) == "windows" else 'c', current_ip_address), shell=True, universal_newlines=True)
        if 'unreachable' in output:
            return False
        else:
            return True
    except Exception:
        return False

def get_date():
    now = datetime.now()  # current date and time
    date_time = now.strftime("%d/%m/%Y, %H:%M:%S")

    return date_time

def add_frame(frame_to_add_to, title_name, array_of_machines, place, sheet, is_carriage):
    if is_carriage:
        padding_x = 0
        label_wraplength = 30
        label_background = GREY
        button_font = 'Helvetica 6 bold'
        row_place = 0
        column_place = place

    else:
        padding_x = 5
        label_wraplength = 0
        label_background = BLACK2
        button_font = 'Helvetica 8 bold'
        row_place =int(place/3)
        column_place =int(place%3)

    frame = tk.Frame(frame_to_add_to, bg=BLACK1, width=TAKASH_WIDTH, height=300)
    frame.grid(row=row_place, column=column_place, padx=padding_x, pady=2, sticky=tk.N)

    name_label = tk.Label(frame, text=title_name,bd=2, bg=label_background,wraplength=label_wraplength, height = 2,fg=WHITE, font='Helvetica 8 bold')
    name_label.grid(row=0, column=0, sticky=tk.NSEW)
    for index, machine in enumerate(array_of_machines):
        if machine.is_working == "עובד":
            background_button_color = GREEN
        else:
            background_button_color = RED

        button = tk.Button(frame, text=machine.who_am_i, bg=background_button_color,
                        fg=WHITE, font=button_font)

        button.configure(command=lambda v=machine.ip,b = button, s =sheet,h = machine.hostname : send_one_ping(v,b,s,h))
        button.grid(row=index + 1, column=0, sticky=tk.EW)

        text_to_add = 'ip: ' + str(machine.ip + "\n" + 'last time: '+ machine.last_send_time)
        Create_tool_tip(button, text = text_to_add, need_delete = False)

def add_data_to_Takash():
    last_takash_name = ""
    for tuple in takash_with_machine:  # tuple of (takash name, machine name)
        if last_takash_name != tuple[0]:
            all_takash[tuple[0]] = Takash(tuple[0])
            last_takash_name = tuple[0]
        all_takash[tuple[0]].add_machine(tuple[1])

def insert_data_to_ui(frame, sheet, is_carriage):
    # first insert the right side of the UI
    path = xl_file

    is_first_time = True
    dict_of_machines_in_sheet = get_data_from_file(path, sheet, is_first_time)

    print("send ping to all of the takashes ip address and save data in xl file at: " + path + "at sheet: " + sheet)
    '''send ping to all of the takashes ip address and save data in xl file'''
    scanning_all(dict_of_machines_in_sheet,sheet)
    '''--------------------------------------------'''
    is_first_time = False

    print("start fetching data from file: " + path + "at sheet: " + sheet)
    dict_of_takashes, dict_of_machines_in_sheet = get_data_from_file(path, sheet, is_first_time)

    temp_machine_name_list = []
    if is_carriage==False:
        # print(len(dict_of_takashes.keys()))
        if len(dict_of_takashes.keys())>9:
            print("לא יכולים להיות יותר מ 9 תק''שים")
            return False

    place = 0
    for key in dict_of_takashes:
        for machine in dict_of_takashes[key].machine_list:
            temp_machine_name_list.append(machine)

        add_frame(frame, key, temp_machine_name_list, place, sheet, is_carriage)
        place = place + 1
        temp_machine_name_list = []
    return True

def get_data_from_file(path, sheet, first_time_flag):

    # To open the workbook, workbook object is created
    workbook_obj = openpyxl.load_workbook(path)
    sheet_obj = workbook_obj[sheet]

    current_takash = sheet_obj.cell(row=2, column=1).value
    takash_with_machine = []  # tuple of (takash name, machine name)
    # Note: The first row or column integer is 1, not 0.

    dict_of_machines_in_sheet = {}
    all_takash = {}
    machines_to_add = 0
    exit_loop_flag = False
    current_row = []
    for row in sheet_obj.iter_rows(min_row=2):
        if exit_loop_flag == True:
            break
        else:
            cell_count = 1
            for cell in row:
                if exit_loop_flag == True:
                    break
                else:
                    if cell_count == 1:
                        if cell.value == None:
                            exit_loop_flag = True
                        else:
                            name_of_takash = cell.value
                            if current_takash != cell.value:
                                current_takash = cell.value
                                machines_to_add = 0
                            machines_to_add += 1
                            current_row = []
                    else:
                        current_row.append(cell.value)
                    cell_count += 1
        dict_of_machines_in_sheet[current_row[2]] = Machine(current_row[2], current_row[1], current_row[0], current_row[4], current_row[3])
        takash_with_machine.append([name_of_takash, dict_of_machines_in_sheet[current_row[2]]])
    takash_with_machine.pop()

    last_takash_name = ""

    if first_time_flag:
        return dict_of_machines_in_sheet

    else:
        for tuple in takash_with_machine:  # tuple of (takash name, machine name)
            if last_takash_name != tuple[0]:
                all_takash[tuple[0]] = Takash(tuple[0])
                last_takash_name = tuple[0]
            all_takash[tuple[0]].add_machine(tuple[1])

        return all_takash, dict_of_machines_in_sheet

class Program(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.run_counter = 0
        self.container = tk.Frame(self, bg=BLACK2, width=right_WIDTH, height=right_HEIGHT)
        self.container.pack(fill=tk.BOTH, expand=True, side=tk.RIGHT)
        self.container.pack_propagate(0)
        # self.container.grid_rowconfigure(0, weight=1)
        # self.container.grid_columnconfigure(0, weight=1)
        self.create_right_side()

    def create_right_side(self):
        self.headquarters_name = self.get_headquarters_name()
        self.top_frame = Top_title(self.container, self.headquarters_name, True)
        self.top_frame.pack(fill=tk.X)

        self.server_frame = Servers_frame(self.container, self)
        self.server_frame.pack()
        self.label_counter =tk.Label(self.container, text= self.run_counter, bg=BLACK2, fg=WHITE, font='Helvetica 18 bold')
        self.label_counter.pack()

    def get_headquarters_name(self):  ### temp for now

        workbook_obj = openpyxl.load_workbook(xl_file)
        sheet_obj = workbook_obj["שרתים"]

        current_takash = sheet_obj.cell(row=1, column=8).value
        return (current_takash)

    def update_data_in_ui(self):
      x = insert_data_to_ui(app.server_frame, "שרתים", False)
      self.run_counter += 1
      self.label_counter.configure(text=self.run_counter)
      self.top_frame.title_top.configure(text="david")
      self.after(5000, self.update_data_in_ui)

class Servers_frame(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, bg=BLACK2)

class Top_title(tk.Frame):
    def __init__(self, parent, title, need_buttons):
        tk.Frame.__init__(self, parent, bg=BLACK2)

        self.upper_frame = tk.Frame(self, bg=BLACK2)
        self.upper_frame.pack(fill=tk.X)

        if need_buttons == True:
            self.setting_button = tk.Button(self.upper_frame, text=" הגדרות", bg=BLACK1, fg=WHITE,
                                            font='Helvetica 8 bold', width=6)
            self.setting_button.grid(row=0, column=0, sticky=tk.W, padx=5)
            self.statistics_button = tk.Button(self.upper_frame, text="סטיסטיקה", bg=BLACK1, fg=WHITE,
                                               font='Helvetica 8 bold', width=6)
            self.statistics_button.grid(row=1, column=0, sticky=tk.W, padx=5)

        self.title_top = tk.Label(self.upper_frame, text=title, bg=BLACK2, fg=WHITE, font='Helvetica 18 bold')
        self.title_top.grid(row=0, column=1, rowspan=2)
        self.upper_frame.grid_rowconfigure(1, weight=1)
        self.upper_frame.grid_columnconfigure(1, weight=1)

        span = tk.Frame(self, bg=BLACK1, height=2)
        span.pack(fill=tk.X, pady=5, padx=10)

class ToolTip(object):

    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 57
        y = y + cy + self.widget.winfo_rooty() +27
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                      background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                      font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

def Create_tool_tip(widget, text, need_delete):
    # if need_delete:
    #     toolTip.widget = None
    #     toolTip.tipwindow = None
    #     toolTip.id = None
    #     toolTip.x = 0
    #     toolTip.y = 0
    toolTip = ToolTip(widget)
    def enter(event):
        toolTip.showtip(text)
    def leave(event):
        toolTip.hidetip()
    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)

def scanning_all(dict_of_all_addresses, sheet):

    path = xl_file
    workbook_obj = openpyxl.load_workbook(path)

    sheet_obj = workbook_obj[sheet]

    list_address = []
    for key in dict_of_all_addresses:
        list_address.append(dict_of_all_addresses[key].ip)

    with concurrent.futures.ThreadPoolExecutor() as executor:
        results = executor.map(send_ping, list_address)

    i = 0
    results = list(results)
    for key in dict_of_all_addresses:
        dict_of_all_addresses[key].is_working = results[i]
        time = get_date()
        dict_of_all_addresses[key].last_send_time = time
        i += 1

    for key in dict_of_all_addresses:
        update_file(sheet_obj, key, dict_of_all_addresses[key].is_working, dict_of_all_addresses[key].last_send_time)

    print("saved all new data to input_data.xlsx")
    workbook_obj.save(path)

def update_file(sheet_obj, hostname, status, time):

    cell_positions = ''
    time_cell_positions=''
    for row in sheet_obj.rows:
        element = row[3]
        if element.value == hostname:
            cell_col_position = 'E'
            cell_row_position = element.row
            cell_positions = cell_col_position + str(cell_row_position)
            time_cell_positions = 'F' + str(cell_row_position)

    if status == True:
        sheet_obj[cell_positions] = "עובד"
        sheet_obj[cell_positions].fill = fill = openpyxl.styles.GradientFill(stop=("80ff72", "7ee8fa"))
        # print( sheet_obj[cell_positions].value)

    else:
        sheet_obj[cell_positions] = "לא עובד"
        # print(sheet_obj[cell_positions].value)
        sheet_obj[cell_positions].fill = fill = openpyxl.styles.GradientFill(stop=("fc9842", "fe5f75"))

    sheet_obj[time_cell_positions] = time


if __name__ == '__main__':
    app = Program()

    x = insert_data_to_ui(app.server_frame, "שרתים", False)
    app.update_data_in_ui()

    app.mainloop()